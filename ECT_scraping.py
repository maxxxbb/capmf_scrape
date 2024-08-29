import requests
from bs4 import BeautifulSoup
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


## 1. Define Input and Output paths


url = 'https://www.energychartertreaty.org/treaty/contracting-parties-and-signatories/'
output_path = r"V:\ENVINFO\BACKUP\STATA\IPAC\CAP\Excel files\1_RawDataCAP\Extension\Energy_charter_treaty_2024.xlsx"


## 1.1 Manually define withdrawal data
withdrawal_data = {
    'Italy': {
        'date_withdrawal_notification': '2014-12-31',
        'date_withdrawal_effect': '2016-01-01',
        'date_ratification': '1997-12-05',
    },
    'France': {
        'date_withdrawal_notification': '2022-12-07',
        'date_withdrawal_effect': '2023-12-08',
        'date_ratification': '1999-09-01'
    },
    'Germany': {
        'date_withdrawal_notification': '2022-12-19',
        'date_withdrawal_effect': '2023-12-20',
        'date_ratification': '1997-03-14',
    },
    'Poland': {
        'date_withdrawal_notification': '2022-12-28',
        'date_withdrawal_effect': '2023-12-29',
        'date_ratification': '2000-11-24',
    },
    'Luxembourg': {
        'date_withdrawal_notification': '2023-06-16',
        'date_withdrawal_effect': '2024-06-17',
    },
    'European Union and Euratom': {
        'date_withdrawal_notification': '2024-05-30',
        'date_withdrawal_effect': '2025-05-30',
    },
    'Spain': {
        'date_withdrawal_notification': '2024-04-16',
        'date_withdrawal_effect': '2025-04-17',
    },
    'United Kingdom': {
        'date_withdrawal_notification': '2024-04-26',
        'date_withdrawal_effect': '2025-04-27',
    }
}
 
## 2. Define Functions 

def fetch_country_links(url):
    """
    Fetches country links from ECT members.
    
    Args:
        url (str): Url of ECT webpage
        
    Returns:
        dict: Dictionary containing country names as keys and their respective URLs as values.
    """
    response = requests.get(url)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, 'html.parser')

    country_links = {}
    for figure in soup.find_all('figure', class_='image'):
        a_tag = figure.find('a')
        if a_tag:
            country_name = a_tag.get('title')
            country_url = a_tag.get('href')
            country_links[country_name] = f'https://www.energychartertreaty.org{country_url}'

    return country_links

def scrape_country_data(country, link):
    """
    Main scraping function: Takes dictionary of country and links an  
    extracts ECT dates (signature, ratification, deposit) for a country.
    
    Args:
        country (str): The name of the country.
        link (str): The URL of the country's page.
        
    Returns:
        list: List containing the country data.
    """
    # send get request to fetch HTML content
    response = requests.get(link)
    response.raise_for_status() # Raises an error when website doesnt respond
    
    # parse the HTML content
    soup = BeautifulSoup(response.text, 'html.parser')

    # find the section containing the data on Energy Charter Treaty
    section = soup.find('strong', string=lambda text: text and '1994 Energy Charter Treaty' in text)
    
    try:
        print(f"Processing country: {country}, URL: {link}")  # Debug print
        if section:
            title = section.get_text(strip=True)
            dates = []
            next_element = section.find_next()
            while next_element and (next_element.name != 'strong'):
                if next_element.name == 'ul':
                    dates.extend([li.get_text(strip=True) for li in next_element.find_all('li')])
                next_element = next_element.find_next()
            while len(dates) < 4:  # fill in missing dates with None if less than 4 dates are found
                dates.append(None)
            print(f"Scraped data for {country}")
            return [country] + [title] + dates ## add country and title to the list
        return []
    
    except Exception as e:
            print(f"An error occurred for country: {country}, URL: {link}")
            print(f"Error: {e}")

def extract_date(text, prefix):
    """
    Extracts the date - in this case year - from a text string based on the given prefix.

    " signed on 12 December 1994" -> 1994 "
    
    Args:
        text (str): The text containing the date.
        prefix (str): The prefix to look for in the text.

    Returns:
        int or None: The extracted year as an integer, or None if not found.
    
    """
    if pd.isnull(text):
        return None
    if prefix in text:
        try:
            return pd.to_datetime(text.split(prefix)[1].strip(), format='%d %B %Y').year
        except (IndexError, ValueError):
            return None
    return None

def update_dataframe_withdrawals(df, withdrawal_data):
    """
    Updates the dataframe with manually extracted withdrawal data.
    
    Args:
        df (pd.DataFrame): Main dataframe.
        withdrawal_data (dict): The dictionary containing withdrawal data.
        
    Returns:
        pd.DataFrame: The updated dataframe.
    """
    
    ## loops over manually created withdrawal dictionary

    for country, dates in withdrawal_data.items():
        if country in df['Country'].values:
            for key, value in dates.items():
                if value != 'N.A.':
                    df.loc[df['Country'] == country, key] = pd.to_datetime(value).year
                else:
                    df.loc[df['Country'] == country, key] = value
        else:
            new_row = {**{'Country': country}, **{k: (pd.to_datetime(v).year if v != 'N.A.' else v) for k, v in dates.items()}}
            df = pd.concat([df, pd.DataFrame(new_row, index=[0])], ignore_index=True)

    return df

def add_readme_to_worksheet(ws):
    """
    Adds a Readme content to the xlsx.
    
    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet to add the Readme content to.
    """

    readme_text = [
        'This file contains the data scraped from the Energy Charter Treaty website: ',
        'https://www.energychartertreaty.org/treaty/contracting-parties-and-signatories/',
        'Countries Italy, France, Germany, and Poland have already effectively withdrawn from the treaty at the time of the data extraction (July 2024), ',
        'so their withdrawal and signature dates did not appear on the website and were manually added.',
        'The pyhton script used to extract the data and create this file can be found at',
        '\\\\main.oecd.org\\ASgenENV\\ENVINFO\\BACKUP\\STATA\\IPAC\\CAP\\Others\\Scraping_Scripts\\ECT_scraping.py'
    ]

    for i, line in enumerate(readme_text, start=1):
        ws.cell(row=i, column=1, value=line)
        if 'https://' in line or '\\\\' in line:
            ws.cell(row=i, column=1).hyperlink = line

def save_to_excel(output_path, df):
    """
    Saves the dataframe and Readme content to an Excel file.
    
    Args:
        output_path (str): The path to save the Excel file to.
        df (pd.DataFrame): The dataframe containing the data.
    """
    wb = Workbook()
    ws_readme = wb.active
    ws_readme.title = "Readme"
    ws_data = wb.create_sheet(title="Energy_charter_treaty_2024")

    add_readme_to_worksheet(ws_readme)

    for row in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(row)

    wb.save(output_path)


#### Run all functions - extracting data and saving to excel

def main():
    
    # get country link dictionary
    country_links = fetch_country_links(url)


    # loop over countries and extract data
    all_data = []
    for country, link in country_links.items():
        data = scrape_country_data(country, link)
        if data:
            all_data.append(data)

    # store in Dataframe and clean a little
    df = pd.DataFrame(all_data)
    df.rename(columns={0: 'Country', 1: 'Title', 2: 'Signature', 3: 'Ratification', 4: 'Deposition', 5: 'Entry into Force', 6: 'Additional'}, inplace=True)

    # extract dates from strings
    df['date_sign'] = df.apply(lambda row: extract_date(row['Signature'], 'signed on'), axis=1)
    df['date_ratification'] = df.apply(lambda row: extract_date(row['Ratification'], 'ratified on'), axis=1)
    df['date_deposit'] = df.apply(lambda row: extract_date(row['Deposition'], 'deposited on'), axis=1)
    df['date_entry_force'] = df.apply(lambda row: extract_date(row['Entry into Force'], 'entered into force on'), axis=1)
    df['date_withdrawal_notification'] = None
    df['date_withdrawal_effect'] = None

    # add withdrawals to dataset

    df = update_dataframe_withdrawals(df, withdrawal_data)
    df.replace({None: np.nan}, inplace=True)

    print("Scraping completed successfully!")

    # save to excel
    save_to_excel(output_path, df)

if __name__ == "__main__":
    main()

